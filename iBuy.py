from splinter import Browser
from openpyxl import load_workbook
import sys,time
import re

def removeHyperlink(link):
	if link.find("HYPERLINK") != -1:
		quoted = re.findall(r'"(.*?)"', link)
		return quoted[1]
	else:
		return link

username = str(sys.argv[1])
password = str(sys.argv[2])
purchaseOrder = str(sys.argv[3])
if len(sys.argv) == 5:
	startRow = str(sys.argv[4])
else:
	startRow = "8"

browser = Browser()
browser.visit("https://eas.admin.uillinois.edu/eas/servlet/EasLogin?redirect=https://webprod.admin.uillinois.edu/ssa/servlet/SelfServiceLogin?appName=edu.uillinois.aits.iBuyHelper")
browser.fill('inputEnterpriseId', username)
browser.fill('password', password)
login = browser.find_by_name('BTN_LOGIN')
login.click()
if browser.is_text_present('non-catalog item', wait_time=15):
	browser.click_link_by_text('non-catalog item')
	wb = load_workbook(purchaseOrder)
	ws = wb["Purchase Request Form"]
	currentRow = startRow
	vendor = ws['A' + str(startRow)].value
	if browser.is_element_present_by_id('ModalPopupIframe', wait_time=5):
		with browser.get_iframe('ModalPopupIframe') as iframe:
			# pause and wait for user to select supplier
			raw_input("You must select the supplier manually. Press Enter when done...")
			while ws['A' + currentRow].value == vendor:
				iframe.fill('NonCatProdDesc', ws['B' + currentRow].value)
				iframe.fill('NonCatCatalogNumber', removeHyperlink(str(ws['C' + currentRow].value)))
				iframe.fill('NonCatQuantity', str(ws['D' + currentRow].value))
				iframe.fill('NonCatUnitPrice', str(ws['E' + currentRow].value))
				iframe.fill('NonCatPkgAmount', "1")
				currentRow = str(int(currentRow) + 1)
				nextItem = iframe.find_by_value("Save and Add Another")
				nextItem.click()
				time.sleep(1)
			closeButton = iframe.find_by_value("Close")
			closeButton.click()
		cart = browser.find_by_id("cartTotal")
		cart.click()
		print 'Script executed successfully. Please review all items before checking out.'
	else:
		print 'Non-Catalog Item page failed to load.'
else:
	print 'Login failed.'
